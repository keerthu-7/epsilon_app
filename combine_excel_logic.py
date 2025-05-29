import openpyxl
from openpyxl.chart import LineChart, Reference, Series
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Font
import os

def get_last_valid_row(ws, col, start_row=4):
    for row in range(ws.max_row, start_row - 1, -1):
        if isinstance(ws.cell(row=row, column=col).value, (int, float)):
            return row
    return start_row

def combine_excel_outputs(source_files, compare_cycles, output_path):
    combined_wb = openpyxl.Workbook()
    combined_ws = combined_wb.active
    start_col = 1

    for file in source_files:
        wb = openpyxl.load_workbook(file)
        ws = wb.active
        file_name = os.path.basename(file).split(".")[0]

        for cycle in compare_cycles:
            col_title = f"{file_name} - Cycle {cycle}"
            for col in range(4):
                combined_ws.cell(row=1, column=start_col + col, value=col_title)

            headers = ['Charging Capacity', 'Charging Voltage', 'Discharging Capacity', 'Discharging Voltage']
            for i, header in enumerate(headers):
                combined_ws.cell(row=3, column=start_col + i, value=header)

            data_start_col = (cycle - 1) * 4 + 1
            max_row = ws.max_row
            data_written = False

            for row in range(4, max_row + 1):
                for offset in range(4):
                    val = ws.cell(row=row, column=data_start_col + offset).value
                    if isinstance(val, (int, float)):
                        combined_ws.cell(row=row, column=start_col + offset, value=val)
                        data_written = True
                    else:
                        combined_ws.cell(row=row, column=start_col + offset, value=None)

            if not data_written:
                print(f"⚠️ No numeric data found for {file_name} - Cycle {cycle}")

            start_col += 4

    # Format headers
    for i in range(1, combined_ws.max_column + 1, 4):
        combined_ws.merge_cells(start_row=1, start_column=i, end_row=1, end_column=i+3)
        combined_ws.merge_cells(start_row=2, start_column=i, end_row=2, end_column=i+3)

        combined_ws.cell(row=1, column=i).alignment = Alignment(horizontal='center', vertical='center')
        combined_ws.cell(row=1, column=i).font = Font(bold=True, size=14)
        combined_ws.cell(row=2, column=i).alignment = Alignment(horizontal='center', vertical='center')
        combined_ws.cell(row=2, column=i).font = Font(bold=False, size=12)

        for j in range(4):
            combined_ws.cell(row=3, column=i + j).alignment = Alignment(horizontal='center')
            combined_ws.cell(row=3, column=i + j).font = Font(bold=True)

    # Add Line Chart
    chart = LineChart()
    chart.title = "Voltage vs Capacity Comparison"
    chart.x_axis.title = "Capacity"
    chart.y_axis.title = "Voltage"
    chart.width = 22
    chart.height = 12

    for i in range(1, combined_ws.max_column + 1, 4):
        name = combined_ws.cell(row=1, column=i).value

        # Charging series
        last_row_chg = get_last_valid_row(combined_ws, i)
        x_vals_chg = Reference(combined_ws, min_col=i, min_row=4, max_row=last_row_chg)
        y_vals_chg = Reference(combined_ws, min_col=i+1, min_row=4, max_row=last_row_chg)
        series_chg = Series(y_vals_chg, title=f"{name} - Charging")
        series_chg.categories = x_vals_chg
        chart.series.append(series_chg)

        # Discharging series
        last_row_dchg = get_last_valid_row(combined_ws, i+2)
        x_vals_dchg = Reference(combined_ws, min_col=i+2, min_row=4, max_row=last_row_dchg)
        y_vals_dchg = Reference(combined_ws, min_col=i+3, min_row=4, max_row=last_row_dchg)
        series_dchg = Series(y_vals_dchg, title=f"{name} - Discharging")
        series_dchg.categories = x_vals_dchg
        chart.series.append(series_dchg)

    chart_position = get_column_letter(combined_ws.max_column + 2) + "4"
    combined_ws.add_chart(chart, chart_position)

    combined_wb.save(output_path)
    print(f"✅ Combined file saved: {output_path}")
