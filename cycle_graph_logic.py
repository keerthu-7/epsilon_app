import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.chart import LineChart, Reference, Series

def generate_graph_file(path, weight, rangec1, rangec2, output_path):
    wb = openpyxl.load_workbook(path)
    shnm = wb.sheetnames

    if len(shnm) < 4:
        raise ValueError("The uploaded Excel file must have at least 4 sheets.")

    ws = wb[shnm[2]]  # Cycle Step Sheet
    ws2 = wb[shnm[3]]  # Data Sheet

    CycStepC, CycStepDC = {}, {}

    # Detect step codes from sheet 3
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
        if row[4].value == "CCCV_DChg":
            CycStepDC[row[1].value] = row[2].value
        elif row[4].value == "CC_Chg":
            CycStepC[row[1].value] = row[2].value

    total_cycles_set = sorted(set(CycStepC.keys()) | set(CycStepDC.keys()))
    total_cycle_count = len(total_cycles_set)

    wbo = openpyxl.Workbook()
    wso = wbo.active
    start_col = 1
    chart = LineChart()
    chart.title = "Combined Voltage vs Capacity"
    chart.x_axis.title = "Capacity"
    chart.y_axis.title = "Voltage"
    chart.width = 20
    chart.height = 10

    for cycle in range(rangec1, rangec2 + 1):
        V_chg, C_chg, V_dchg, C_dchg = [], [], [], []

        for row in ws2.iter_rows():
            if row[3].value == cycle:
                # Charging
                if cycle in CycStepC and row[4].value == CycStepC[cycle]:
                    V_chg.append(row[6].value)
                    C_chg.append(float(row[7].value) / weight)
                # Discharging
                if cycle in CycStepDC and row[4].value == CycStepDC[cycle]:
                    V_dchg.append(row[6].value)
                    C_dchg.append(float(row[7].value) / weight)

        # Write headers
        wso.cell(row=1, column=start_col, value=f"Cycle {cycle}")
        wso.cell(row=2, column=start_col, value="Capacity (Charging)")
        wso.cell(row=2, column=start_col + 1, value="Voltage (Charging)")
        wso.cell(row=2, column=start_col + 2, value="Capacity (Discharging)")
        wso.cell(row=2, column=start_col + 3, value="Voltage (Discharging)")

        # Write charging values
        for i, (cap, volt) in enumerate(zip(C_chg, V_chg), start=3):
            wso.cell(row=i, column=start_col, value=cap)
            wso.cell(row=i, column=start_col + 1, value=volt)

        # Write discharging values
        for j, (cap2, volt2) in enumerate(zip(C_dchg, V_dchg), start=3):
            wso.cell(row=j, column=start_col + 2, value=cap2)
            wso.cell(row=j, column=start_col + 3, value=volt2)

        # Get max rows for chart references
        max_row = max(len(C_chg), len(C_dchg)) + 2  # Account for header rows

        # Chart series
        if V_chg:
            chg_series = Reference(wso, min_col=start_col + 1, min_row=2, max_row=max_row)
            chart.series.append(Series(chg_series, title=f"Charging Cycle {cycle}"))

        if V_dchg:
            dchg_series = Reference(wso, min_col=start_col + 3, min_row=2, max_row=max_row)
            chart.series.append(Series(dchg_series, title=f"Discharging Cycle {cycle}"))

        start_col += 4

    chart_anchor = get_column_letter(start_col + 1) + "5"
    wso.add_chart(chart, chart_anchor)
    wbo.save(output_path)

    return total_cycle_count
