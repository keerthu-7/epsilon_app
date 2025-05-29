from flask import Flask, render_template, request, send_file, redirect, url_for
import os
import tempfile
import openpyxl
from cycle_graph_logic import generate_graph_file
from combine_excel_logic import combine_excel_outputs

app = Flask(__name__)
app.config['UPLOAD_FOLDER'] = tempfile.gettempdir()

@app.route("/")
def index():
    return render_template("index.html")

@app.route("/extract", methods=["POST"])
def extract():
    try:
        excel_file = request.files.get("excel_file")
        weight = float(request.form.get("weight"))
        cycle_start = int(request.form.get("cycle_start"))
        cycle_end = int(request.form.get("cycle_end"))
        filename_output = request.form.get("filename_output") + ".xlsx"

        if not excel_file:
            return "No file uploaded", 400

        temp_dir = app.config['UPLOAD_FOLDER']
        input_path = os.path.join(temp_dir, excel_file.filename)
        output_path = os.path.join(temp_dir, filename_output)

        excel_file.save(input_path)
        total_cycles = generate_graph_file(input_path, weight, cycle_start, cycle_end, output_path)

        return render_template("index.html", download_link=f"/download/{filename_output}", total_cycles=total_cycles)

    except Exception as e:
        return f"<h4>❌ Internal Server Error</h4><p>{str(e)}</p>", 500

@app.route("/download/<filename>")
def download_file(filename):
    file_path = os.path.join(app.config['UPLOAD_FOLDER'], filename)
    if os.path.exists(file_path):
        return send_file(file_path, as_attachment=True)
    else:
        return "File not found.", 404

@app.route("/combine", methods=["POST"])
def combine():
    try:
        uploaded_files = request.files.getlist("excel_files")
        cycle_text = request.form.getlist("comparison_cycles")
        filename_output = request.form.get("combine_filename")

        if not uploaded_files or not cycle_text:
            return "No files or cycles provided", 400

        selected_cycles = [int(c) for c in cycle_text]

        temp_dir = tempfile.mkdtemp()
        file_paths = []
        for file in uploaded_files:
            path = os.path.join(temp_dir, file.filename)
            file.save(path)
            file_paths.append(path)

        output_path = os.path.join(temp_dir, f"{filename_output}.xlsx")
        combine_excel_outputs(file_paths, selected_cycles, output_path)

        return send_file(output_path, as_attachment=True)

    except Exception as e:
        return f"<h4>❌ Internal Error</h4><p>{str(e)}</p>", 500

@app.route("/get_common_cycles", methods=["POST"])
def get_common_cycles():
    try:
        uploaded_files = request.files.getlist("excel_files")
        if not uploaded_files:
            return {"error": "No files uploaded"}, 400

        cycle_sets = []
        for file in uploaded_files:
            wb = openpyxl.load_workbook(file)
            ws = wb.active

            header_cycles = set()
            for col in range(1, ws.max_column + 1, 4):  # Step over each cycle's first column
                val = ws.cell(row=1, column=col).value
                if isinstance(val, str) and val.lower().startswith("cycle"):
                    try:
                        num = int(val.split()[-1])
                        header_cycles.add(num)
                    except ValueError:
                        continue

            cycle_sets.append(header_cycles)

        common_cycles = sorted(list(set.intersection(*cycle_sets)))

        # Debug logs
        print("Uploaded Files:", [f.filename for f in uploaded_files])
        print("Cycle Sets:", cycle_sets)
        print("Common Cycles:", common_cycles)

        return {"common_cycles": common_cycles}

    except Exception as e:
        return {"error": str(e)}, 500

if __name__ == "__main__":
    app.run(debug=True)
